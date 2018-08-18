import math
import openpyxl
import xlrd


def reprint_lognorm(write_work, sheet, write_sheet, dist_col_num, instance_len):
    dist_col = sheet.col(dist_col_num)

    for i in range(len(dist_col)):
        if i % (instance_len + 1) == 0:
            name = dist_col[i].value
            if name == 'END':
                break
        else:
            if i % (instance_len + 1) == 1:
                continue
            else:
                dist_par = dist_col[i].value
                if dist_par != '':
                    try:
                        mean, sig = dist_par.split(', ')
                        mean = float(mean)
                        sig = float(sig)
                    except:
                        continue
                    new_mean, new_sig = convert(mean,sig)
                    new_print = str(new_mean) + ', ' + str(new_sig)
                    write_sheet.cell(row=i+1, column=dist_col_num+1).value = new_print
    write_work.save('sheets/input/tests/FR_Input_st_small_log_sigma.xlsx')


def reprint_triag(write_work, sheet, write_sheet, dist_col_num, instance_len):
    dist_col = sheet.col(dist_col_num)

    for i in range(len(dist_col)):
        if i % (instance_len + 1) == 0:
            name = dist_col[i].value
            if name == 'END':
                break
        else:
            if i % (instance_len + 1) == 1:
                continue
            else:
                dist_par = dist_col[i].value
                if dist_par != '':
                    print('hello?')
                    try:
                        mean, sig = dist_par.split(', ')
                        mean = float(mean)
                        sig = float(sig)
                    except:
                        continue
                    c = mean
                    a = mean - sig
                    b = mean + sig
                    print(a,b,c)
                    new_print = str(a) + ', ' + str(b) + ', ' + str(c)
                    write_sheet.cell(row=i+1, column=dist_col_num+1).value = new_print
    write_work.save('sheets/input/tests/FR_Input_triangle.xlsx')



def convert(mean, sig):

    new_mean = math.exp(mean+(.5 * math.pow(sig,2)))
    new_sig = math.exp(2*mean + math.pow(sig,2))*(math.exp(math.pow(sig,2))-1)

    return new_mean, new_sig



f_len = 11  # number of inputs per fish
zo_len = 11  # number of inputs per zoo
ph_len = 4  # number of inputs per phyto
reg_len = 10  # number of inputs per region
chem_len = 7

all_sheets = xlrd.open_workbook('sheets/input/tests/approach_mean_write.xlsx')
write_work = openpyxl.load_workbook('sheets/input/tests/FR_Input_triangle.xlsx')
reg_sheet_read = all_sheets.sheet_by_index(1)
reg_sheet_write = write_work.worksheets[1]

chem_sheet_read = all_sheets.sheet_by_index(2)
chem_sheet_write = write_work.worksheets[2]

an_sheet_read = all_sheets.sheet_by_index(3)
an_sheet_write = write_work.worksheets[3]


reprint_triag(write_work, reg_sheet_read, reg_sheet_write,2,reg_len)
reprint_triag(write_work, chem_sheet_read, chem_sheet_write, 2, chem_len)
reprint_triag(write_work, an_sheet_read, an_sheet_write, 2, f_len)
reprint_triag(write_work, an_sheet_read, an_sheet_write, 5, zo_len)
reprint_triag(write_work, an_sheet_read, an_sheet_write, 9, ph_len)